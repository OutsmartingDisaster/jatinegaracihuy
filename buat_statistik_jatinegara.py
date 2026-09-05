# -*- coding: utf-8 -*-
"""Buat file Excel statistik kelurahan Kecamatan Jatinegara."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

rows = [
    # (Kelurahan, KodePos, Luas_km2, RW, RT, KK, L, P, Balita, Anak, Remaja, Lansia)
    ("Bali Mester",          "13310", 0.67,  6,  73,  4_202,  5_690,  5_596,   967,   738,  1_739,   793),
    ("Kampung Melayu",       "13320", 0.48,  9, 106, 10_408, 15_892, 15_422, 1_907, 2_681,  5_331, 3_415),
    ("Bidara Cina",          "13330", 1.26, 16, 188, 15_048, 21_072, 21_164, 2_225, 3_026,  3_242, 5_529),
    ("Cipinang Cempedak",    "13340", 1.29, 11, 154, 13_035, 18_909, 19_246, 1_874, 7_460,  6_021, 5_438),
    ("Rawa Bunga",           "13350", 0.88,  9, 109,  8_935, 13_080, 13_075, 1_956, 6_920, 19_420, 3_292),
    ("Cipinang Besar Utara", "13410", 1.15, 14, 177, 18_486, 29_217, 28_193, 3_200, 25_710, 27_100, 1_400),
    ("Cipinang Besar Selatan","13410",1.63, 10, 128, 14_726, 22_586, 22_316, 3_643, 4_566,  4_120, 4_186),
    ("Cipinang Muara",       "13420", 2.90, 16, 176, 22_307, 33_691, 33_671, 3_552, 5_361,  5_630, 7_865),
]

wb = Workbook()
ws = wb.active
ws.title = "Statistik Kelurahan"

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1F4E78")
total_fill = PatternFill("solid", fgColor="D9E1F2")
thin = Side(style="thin", color="999999")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center")

title = "STATISTIK KELURAHAN KECAMATAN JATINEGARA - KOTA ADMINISTRASI JAKARTA TIMUR"
ws.merge_cells("A1:M1")
ws["A1"] = title
ws["A1"].font = Font(bold=True, size=13)
ws["A1"].alignment = center
ws.merge_cells("A2:M2")
ws["A2"] = "Sumber kependudukan: timur.jakarta.go.id | Luas wilayah: BPS"
ws["A2"].font = Font(italic=True, size=9, color="555555")
ws["A2"].alignment = center

headers = ["No", "Kelurahan", "Kode Pos", "Luas (km2)", "RW", "RT", "KK",
           "Laki-laki", "Perempuan", "Total Jiwa", "Balita", "Anak", "Remaja", "Lansia"]
# Note: 14 columns A..N; fix title merges accordingly
ws.unmerge_cells("A1:M1"); ws.unmerge_cells("A2:M2")
ws.merge_cells("A1:N1"); ws.merge_cells("A2:N2")

hdr_row = 3
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=hdr_row, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

r = hdr_row + 1
for i, (kel, pos, luas, rw, rt, kk, l, p, balita, anak, remaja, lansia) in enumerate(rows, 1):
    total = l + p
    vals = [i, kel, pos, luas, rw, rt, kk, l, p, total, balita, anak, remaja, lansia]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.border = border
        if c != 2:
            cell.alignment = Alignment(horizontal="center")
    r += 1

# Total row
totals = ["", "TOTAL", "",
          round(sum(x[2] for x in rows), 2),
          sum(x[3] for x in rows), sum(x[4] for x in rows), sum(x[5] for x in rows),
          sum(x[6] for x in rows), sum(x[7] for x in rows),
          sum(x[6] + x[7] for x in rows),
          sum(x[8] for x in rows), sum(x[9] for x in rows),
          sum(x[10] for x in rows), sum(x[11] for x in rows)]
for c, v in enumerate(totals, 1):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(bold=True)
    cell.fill = total_fill
    cell.border = border
    if c != 2:
        cell.alignment = Alignment(horizontal="center")

# Column widths
widths = [4, 22, 9, 11, 6, 7, 9, 11, 11, 12, 9, 9, 9, 9]
for c, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(c)].width = w
ws.freeze_panes = "A4"

out = r"C:\Users\Rio\Downloads\101-sesi5\statistik_kelurahan_jatinegara.xlsx"
wb.save(out)
print("saved:", out)
